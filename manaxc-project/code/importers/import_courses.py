#!/usr/bin/env python3
"""
Import courses from Course Rating Logic.xlsx to Supabase.

This script imports course data including mile_difficulty ratings
from the Course Rating Logic Excel file.
"""

import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl
from decimal import Decimal
import argparse

# Load environment variables
load_dotenv()

# Supabase configuration
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
COURSE_RATING_FILE = os.getenv('COURSE_RATING_FILE', '../../reference/data/Course Rating Logic.xlsx')

# Initialize Supabase client
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


def parse_course_name_location(course_str):
    """
    Parse course string like 'Crystal Springs | 2.95 miles'
    into name, location, and distance.

    Returns: (name, location, distance_display)
    """
    if '|' not in course_str:
        return course_str, None, None

    parts = course_str.split('|')
    name = parts[0].strip()
    distance_display = parts[1].strip() if len(parts) > 1 else None

    # For location, we'll need to look it up or derive from the name
    # For now, return None and we'll fill it manually
    location = None

    return name, location, distance_display


def import_courses(limit=None, dry_run=False):
    """
    Import courses from Course Rating Logic Excel file.

    Args:
        limit: Import only first N rows (for testing)
        dry_run: Don't actually insert, just print what would be imported
    """
    # Get the script directory and build absolute path
    script_dir = Path(__file__).parent
    excel_path = script_dir / COURSE_RATING_FILE

    if not excel_path.exists():
        print(f"❌ Error: File not found at {excel_path}")
        sys.exit(1)

    print(f"📂 Reading courses from: {excel_path}")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active  # First sheet contains the course data

    courses = []
    errors = []

    # Read header row (row 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

    print(f"📋 Found columns: {list(headers.keys())}")

    # Map column names
    course_col = headers.get('Race Course')
    miles_col = headers.get('Distance in Miles')
    meters_col = headers.get('Distance In Meters')
    difficulty_col = headers.get('mile_difficulty')
    description_col = headers.get('Description')

    if not all([course_col, miles_col, meters_col, difficulty_col]):
        print("❌ Error: Missing required columns")
        print(f"   Expected: Race Course, Distance in Miles, Distance In Meters, mile_difficulty")
        print(f"   Found: {list(headers.keys())}")
        sys.exit(1)

    # Read data rows (skip header)
    row_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if limit and row_count >= limit:
            break

        # Get values
        course_str = row[course_col - 1] if course_col else None
        distance_miles = row[miles_col - 1] if miles_col else None
        distance_meters = row[meters_col - 1] if meters_col else None
        mile_difficulty = row[difficulty_col - 1] if difficulty_col else None
        description = row[description_col - 1] if description_col else None

        # Skip empty rows
        if not course_str:
            continue

        # Parse course name
        name, location, distance_display = parse_course_name_location(course_str)

        # Build course record
        course = {
            'name': name,
            'location': location or 'TBD',  # We'll update this manually
            'distance_meters': int(distance_meters) if distance_meters else None,
            'distance_display': distance_display,
            'difficulty_rating': float(mile_difficulty) if mile_difficulty else None,
            'terrain_description': description
        }

        courses.append(course)
        row_count += 1

        if dry_run:
            print(f"  {row_count}. {name}")
            print(f"     Distance: {distance_meters}m ({distance_display})")
            print(f"     Difficulty: {mile_difficulty}")
            print(f"     Description: {description}")
            print()

    print(f"\n📊 Found {len(courses)} courses")

    if dry_run:
        print("\n🔍 DRY RUN - No data inserted")
        return

    # Insert courses
    print("\n💾 Inserting courses into Supabase...")
    inserted = 0
    skipped = 0

    for idx, course in enumerate(courses, 1):
        try:
            # Check if course already exists
            existing = supabase.table('courses').select('id').eq('name', course['name']).execute()

            if existing.data:
                print(f"  ⏭️  {idx}. {course['name']} - already exists, skipping")
                skipped += 1
                continue

            # Insert course
            result = supabase.table('courses').insert(course).execute()
            print(f"  ✅ {idx}. {course['name']} - imported")
            inserted += 1

        except Exception as e:
            error_msg = f"Row {idx}: {course['name']} - {str(e)}"
            errors.append(error_msg)
            print(f"  ❌ {idx}. {error_msg}")

    # Summary
    print("\n" + "="*60)
    print(f"✅ Import complete!")
    print(f"   Inserted: {inserted}")
    print(f"   Skipped:  {skipped}")
    print(f"   Errors:   {len(errors)}")

    if errors:
        print("\n⚠️  Errors encountered:")
        for error in errors:
            print(f"   - {error}")

    return inserted, skipped, errors


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Import courses from Excel to Supabase')
    parser.add_argument('--limit', type=int, help='Import only first N rows (for testing)')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be imported without inserting')

    args = parser.parse_args()

    print("🚀 ManaXC Course Importer")
    print("="*60)

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("❌ Error: SUPABASE_URL and SUPABASE_KEY must be set in .env file")
        print("   Copy .env.example to .env and add your credentials")
        sys.exit(1)

    import_courses(limit=args.limit, dry_run=args.dry_run)
