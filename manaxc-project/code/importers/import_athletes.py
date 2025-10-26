#!/usr/bin/env python3
"""
Import athletes from westmont-xc-results.xlsx to Supabase.

This script imports athlete data from the Master Athletes worksheet,
using the nameparser library to handle complex names.
"""

import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl
from nameparser import HumanName
import argparse

# Load environment variables
load_dotenv()

# Supabase configuration
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH', '../../reference/data/westmont-xc-results.xlsx')

# Initialize Supabase client
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


def parse_athlete_name(full_name):
    """
    Parse full name into first and last name using nameparser library.

    Handles complex names like:
    - "John Smith"
    - "Juan de la Vega Jr."
    - "Sean O'Connor"
    - "Smith, John" (Last, First format)

    Returns: (first_name, last_name, middle_name, suffix)
    """
    if not full_name:
        return None, None, None, None

    name = HumanName(full_name)

    return (
        name.first or None,
        name.last or None,
        name.middle or None,
        name.suffix or None
    )


def import_athletes(limit=None, dry_run=False):
    """
    Import athletes from Master Athletes worksheet.

    Args:
        limit: Import only first N rows (for testing)
        dry_run: Don't actually insert, just print what would be imported
    """
    # Get the script directory and build absolute path
    script_dir = Path(__file__).parent
    excel_path = script_dir / EXCEL_FILE_PATH

    if not excel_path.exists():
        print(f"❌ Error: File not found at {excel_path}")
        sys.exit(1)

    print(f"📂 Reading athletes from: {excel_path}")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Find Master Athletes sheet
    if 'Master Athletes' not in wb.sheetnames:
        print(f"❌ Error: 'Master Athletes' worksheet not found")
        print(f"   Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb['Master Athletes']
    print(f"📋 Reading from 'Master Athletes' worksheet")

    athletes = []
    errors = []

    # Read header row (row 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

    print(f"📋 Found columns: {list(headers.keys())}")

    # Map column names (adjust based on actual Excel structure)
    # We'll need to see the actual column names first
    name_col = headers.get('Name') or headers.get('Athlete Name') or headers.get('Full Name')
    grad_year_col = headers.get('Grad Year') or headers.get('Graduation Year') or headers.get('Year')
    gender_col = headers.get('Gender') or headers.get('Sex')
    school_col = headers.get('School')

    if not name_col:
        print("❌ Error: Could not find name column")
        print(f"   Available columns: {list(headers.keys())}")
        print("\n💡 Please check the Excel file and update column mapping")
        sys.exit(1)

    print(f"\n📍 Using columns:")
    print(f"   Name: {name_col}")
    print(f"   Grad Year: {grad_year_col}")
    print(f"   Gender: {gender_col}")
    print(f"   School: {school_col}")

    # First, get or create Westmont High School
    westmont = supabase.table('schools').select('id').eq('name', 'Westmont High School').execute()
    if not westmont.data:
        print("❌ Error: Westmont High School not found in database")
        print("   Please run the database setup script first")
        sys.exit(1)

    westmont_id = westmont.data[0]['id']
    print(f"✅ Found Westmont High School (ID: {westmont_id})")

    # Read data rows (skip header)
    row_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if limit and row_count >= limit:
            break

        # Get values
        full_name = row[name_col - 1] if name_col else None
        grad_year = row[grad_year_col - 1] if grad_year_col else None
        gender = row[gender_col - 1] if gender_col else None

        # Skip empty rows
        if not full_name:
            continue

        # Parse name
        first_name, last_name, middle_name, suffix = parse_athlete_name(full_name)

        if not first_name or not last_name:
            errors.append(f"Row {row_idx}: Could not parse name '{full_name}'")
            continue

        # Normalize gender
        if gender:
            gender = gender.strip().upper()
            if gender in ['M', 'MALE', 'BOY', 'BOYS']:
                gender = 'M'
            elif gender in ['F', 'FEMALE', 'GIRL', 'GIRLS']:
                gender = 'F'

        # Build athlete record
        athlete = {
            'first_name': first_name,
            'last_name': last_name,
            'middle_name': middle_name,
            'suffix': suffix,
            'grad_year': int(grad_year) if grad_year else None,
            'gender': gender,
            'school_id': westmont_id
        }

        athletes.append(athlete)
        row_count += 1

        if dry_run:
            print(f"  {row_count}. {first_name} {last_name}")
            if middle_name:
                print(f"     Middle: {middle_name}")
            if suffix:
                print(f"     Suffix: {suffix}")
            print(f"     Grad: {grad_year}, Gender: {gender}")
            print()

    print(f"\n📊 Found {len(athletes)} athletes")

    if dry_run:
        print("\n🔍 DRY RUN - No data inserted")
        return

    # Insert athletes
    print("\n💾 Inserting athletes into Supabase...")
    inserted = 0
    skipped = 0

    for idx, athlete in enumerate(athletes, 1):
        try:
            # Check if athlete already exists (by name + grad year)
            existing = supabase.table('athletes')\
                .select('id')\
                .eq('first_name', athlete['first_name'])\
                .eq('last_name', athlete['last_name'])\
                .eq('grad_year', athlete['grad_year'])\
                .execute()

            if existing.data:
                print(f"  ⏭️  {idx}. {athlete['first_name']} {athlete['last_name']} ({athlete['grad_year']}) - already exists, skipping")
                skipped += 1
                continue

            # Insert athlete
            result = supabase.table('athletes').insert(athlete).execute()
            print(f"  ✅ {idx}. {athlete['first_name']} {athlete['last_name']} ({athlete['grad_year']}) - imported")
            inserted += 1

        except Exception as e:
            error_msg = f"Row {idx}: {athlete['first_name']} {athlete['last_name']} - {str(e)}"
            errors.append(error_msg)
            print(f"  ❌ {idx}. {error_msg}")

    # Summary
    print("\n" + "="*60)
    print(f"✅ Import complete!")
    print(f"   Inserted: {inserted}")
    print(f"   Skipped:  {skipped}")
    print(f"   Errors:   {len(errors)}")

    if errors:
        print("\n⚠️  Errors encountered:")
        for error in errors[:10]:  # Show first 10 errors
            print(f"   - {error}")
        if len(errors) > 10:
            print(f"   ... and {len(errors) - 10} more errors")

    return inserted, skipped, errors


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Import athletes from Excel to Supabase')
    parser.add_argument('--limit', type=int, help='Import only first N rows (for testing)')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be imported without inserting')

    args = parser.parse_args()

    print("🚀 ManaXC Athlete Importer")
    print("="*60)

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("❌ Error: SUPABASE_URL and SUPABASE_KEY must be set in .env file")
        print("   Copy .env.example to .env and add your credentials")
        sys.exit(1)

    import_athletes(limit=args.limit, dry_run=args.dry_run)
