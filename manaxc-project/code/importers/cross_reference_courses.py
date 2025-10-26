#!/usr/bin/env python3
"""
Cross-reference courses from Westmont Excel with Course Rating Logic.

This will show us:
1. Which Westmont courses have difficulty ratings
2. Which Westmont courses are missing difficulty ratings
3. Which rated courses aren't in Westmont data
"""
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import re

def parse_distance_to_meters(distance_str):
    """Convert distance string to meters"""
    if not distance_str:
        return None

    distance_str = distance_str.strip().lower()
    match = re.search(r'(\d+\.?\d*)', distance_str)
    if not match:
        return None

    value = float(match.group(1))

    if 'mile' in distance_str:
        return int(value * 1609.34)
    elif 'k' in distance_str or 'km' in distance_str:
        return int(value * 1000)
    elif 'm' in distance_str:
        return int(value)
    else:
        return int(value * 1609.34)

def parse_course_string(course_str):
    """Parse course string like 'Crystal Springs | 2.95 Miles'"""
    if not course_str:
        return None, None, None

    if '|' in course_str:
        parts = course_str.split('|')
        name = parts[0].strip()
        distance_display = parts[1].strip() if len(parts) > 1 else None
    else:
        name = course_str.strip()
        distance_display = None

    distance_meters = parse_distance_to_meters(distance_display) if distance_display else None
    return name, distance_display, distance_meters

def extract_westmont_courses():
    """Extract courses from Westmont Excel"""
    print("📂 Reading Westmont XC Results...")
    excel_file = '/Users/ron/manaxc/manaxc-project/reference/data/westmont-xc-results.xlsx'

    z = zipfile.ZipFile(excel_file)

    # Read shared strings
    strings_xml = z.read('xl/sharedStrings.xml')
    strings_root = ET.fromstring(strings_xml)
    ns_strings = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    shared_strings = []
    for si in strings_root.findall('.//s:si', ns_strings):
        t = si.find('.//s:t', ns_strings)
        if t is not None and t.text:
            shared_strings.append(t.text)

    # Read worksheet
    sheet_xml = z.read('xl/worksheets/sheet7.xml')
    sheet_root = ET.fromstring(sheet_xml)
    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

    # Parse courses
    rows = sheet_root.findall('.//s:row', ns)
    course_names = set()

    for row in rows[1:]:
        cells = row.findall('.//s:c', ns)
        if len(cells) < 6:
            continue

        course_cell = cells[5] if len(cells) > 5 else None
        if course_cell is None:
            continue

        v = course_cell.find('.//s:v', ns)
        t = course_cell.get('t')

        if v is not None and t == 's':
            idx_str = int(v.text)
            if idx_str < len(shared_strings):
                course_name = shared_strings[idx_str]
                if course_name:
                    course_names.add(course_name)

    # Parse into structured data
    courses = {}
    for course_str in course_names:
        name, distance_display, distance_meters = parse_course_string(course_str)
        if name:
            key = f"{name}|{distance_meters}" if distance_meters else name
            courses[key] = {
                'name': name,
                'distance_display': distance_display,
                'distance_meters': distance_meters,
                'original': course_str
            }

    print(f"   ✅ Found {len(courses)} unique courses")
    return courses

def extract_rated_courses():
    """Extract courses with difficulty ratings from Course Rating Logic"""
    print("\n📂 Reading Course Rating Logic...")
    excel_file = '/Users/ron/manaxc/manaxc-project/reference/data/Course Rating Logic.xlsx'

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active

    # Read headers
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

    course_col = headers.get('Race Course')
    miles_col = headers.get('Distance in Miles')
    meters_col = headers.get('Distance In Meters')
    difficulty_col = headers.get('mile_difficulty')

    # Read courses
    courses = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        course_str = row[course_col - 1] if course_col else None
        distance_meters = row[meters_col - 1] if meters_col else None
        difficulty = row[difficulty_col - 1] if difficulty_col else None

        if not course_str:
            continue

        name, distance_display, parsed_meters = parse_course_string(course_str)

        # Use provided meters or parsed meters
        meters = int(distance_meters) if distance_meters else parsed_meters

        if name:
            key = f"{name}|{meters}" if meters else name
            courses[key] = {
                'name': name,
                'distance_display': distance_display,
                'distance_meters': meters,
                'difficulty_rating': float(difficulty) if difficulty else None,
                'original': course_str
            }

    print(f"   ✅ Found {len(courses)} rated courses")
    return courses

def cross_reference(westmont_courses, rated_courses):
    """Cross-reference the two course lists"""
    print("\n" + "=" * 100)
    print("CROSS-REFERENCE ANALYSIS")
    print("=" * 100)

    # Find matches
    matched = []
    westmont_only = []

    for key, course in westmont_courses.items():
        if key in rated_courses:
            matched.append({
                'westmont': course,
                'rated': rated_courses[key]
            })
        else:
            westmont_only.append(course)

    # Find rated courses not in Westmont data
    rated_only = []
    for key, course in rated_courses.items():
        if key not in westmont_courses:
            rated_only.append(course)

    # Display results
    print(f"\n📊 SUMMARY")
    print(f"   Total Westmont courses:      {len(westmont_courses)}")
    print(f"   Total rated courses:         {len(rated_courses)}")
    print(f"   ✅ Matched (with ratings):   {len(matched)}")
    print(f"   ⚠️  Westmont only (no rating): {len(westmont_only)}")
    print(f"   📋 Rated only (not in data):  {len(rated_only)}")

    # Show matched courses with ratings
    if matched:
        print(f"\n" + "=" * 100)
        print(f"✅ MATCHED COURSES WITH DIFFICULTY RATINGS ({len(matched)})")
        print("=" * 100)
        print(f"\n{'Course Name':<40} {'Distance':<15} {'Difficulty':<12}")
        print("-" * 100)
        for item in sorted(matched, key=lambda x: x['westmont']['name']):
            name = item['westmont']['name'][:38]
            dist = item['westmont']['distance_display'] or 'N/A'
            diff = item['rated']['difficulty_rating']
            diff_str = f"{diff:.2f}" if diff else "N/A"
            print(f"{name:<40} {dist:<15} {diff_str:<12}")

    # Show Westmont courses missing ratings
    if westmont_only:
        print(f"\n" + "=" * 100)
        print(f"⚠️  WESTMONT COURSES WITHOUT DIFFICULTY RATINGS ({len(westmont_only)})")
        print("=" * 100)
        print(f"\n{'Course Name':<40} {'Distance':<15}")
        print("-" * 100)
        for course in sorted(westmont_only, key=lambda x: x['name']):
            name = course['name'][:38]
            dist = course['distance_display'] or 'N/A'
            print(f"{name:<40} {dist:<15}")

    # Show rated courses not used in Westmont data
    if rated_only:
        print(f"\n" + "=" * 100)
        print(f"📋 RATED COURSES NOT IN WESTMONT DATA ({len(rated_only)})")
        print("=" * 100)
        print(f"\n{'Course Name':<40} {'Distance':<15} {'Difficulty':<12}")
        print("-" * 100)
        for course in sorted(rated_only, key=lambda x: x['name'])[:20]:
            name = course['name'][:38]
            dist = course['distance_display'] or 'N/A'
            diff = course['difficulty_rating']
            diff_str = f"{diff:.2f}" if diff else "N/A"
            print(f"{name:<40} {dist:<15} {diff_str:<12}")

        if len(rated_only) > 20:
            print(f"\n... and {len(rated_only) - 20} more rated courses not in Westmont data")

    return matched, westmont_only, rated_only

def main():
    """Run cross-reference analysis"""
    print("=" * 100)
    print("COURSE CROSS-REFERENCE TOOL")
    print("=" * 100)

    # Extract courses from both sources
    westmont_courses = extract_westmont_courses()
    rated_courses = extract_rated_courses()

    # Cross-reference
    matched, westmont_only, rated_only = cross_reference(westmont_courses, rated_courses)

    # Recommendations
    print("\n" + "=" * 100)
    print("💡 RECOMMENDATIONS")
    print("=" * 100)

    if matched:
        print(f"\n✅ {len(matched)} courses have difficulty ratings and can be imported with ratings")

    if westmont_only:
        print(f"\n⚠️  {len(westmont_only)} Westmont courses are missing difficulty ratings:")
        print("   - These can still be imported without difficulty ratings")
        print("   - Ratings can be added manually later")
        print("   - Or we can calculate ratings based on historical performance data")

    if rated_only:
        print(f"\n📋 {len(rated_only)} rated courses aren't used in Westmont historical data")
        print("   - These are available for future meets")
        print("   - Can be imported now for completeness")

if __name__ == '__main__':
    main()
